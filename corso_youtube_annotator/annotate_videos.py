#!/usr/bin/env python3
"""
One-row-at-a-time YouTube → Ollama vision → draft Excel annotations.

Switch models: change LOCAL_MODEL_NAME (and FALLBACK_MODELS) in the config block below.
Requires: yt-dlp and ffmpeg on PATH; Ollama running with a vision model pulled.
"""

from __future__ import annotations

import base64
import json
import logging
import os
import re
import shutil
import subprocess
import sys
import time
import tempfile
from pathlib import Path
from typing import Any

try:
    from dotenv import load_dotenv

    load_dotenv(Path(__file__).resolve().parent / ".env")
except ImportError:
    pass

# yt-dlp invokes node/deno for YouTube "n challenge"; minimal PATH breaks subprocess.
if sys.platform == "darwin":
    _brew_paths = ("/opt/homebrew/bin", "/usr/local/bin")
elif sys.platform.startswith("linux"):
    _brew_paths = ("/usr/local/bin",)
else:
    _brew_paths = ()
_path = os.environ.get("PATH", "")
for _p in _brew_paths:
    if _p not in _path.split(os.pathsep):
        os.environ["PATH"] = _p + os.pathsep + _path
        _path = os.environ["PATH"]

# -----------------------------------------------------------------------------
# Config (edit here or override with environment variables)
# -----------------------------------------------------------------------------
INPUT_XLSX = os.environ.get(
    "INPUT_XLSX",
    str(
        Path(__file__).resolve().parent.parent
        / "docs"
        / "Corso Group YouTube Videos (Random Split 17) - Santosh.xlsx"
    ),
)
OUTPUT_XLSX = os.environ.get(
    "OUTPUT_XLSX",
    str(
        Path(__file__).resolve().parent
        / "Corso Group YouTube Videos (Random Split 17) - Santosh_annotated_draft.xlsx"
    ),
)
SHEET_NAME: str | None = os.environ.get("SHEET_NAME") or "youtube_videos"
YOUTUBE_ID_COLUMN: str | None = os.environ.get("YOUTUBE_ID_COLUMN") or None
YOUTUBE_LINK_COLUMN: str | None = os.environ.get("YOUTUBE_LINK_COLUMN") or None

FRAME_INTERVAL_SEC = float(os.environ.get("FRAME_INTERVAL_SEC", "8"))
# Cap on frames sent to vision; 0 = no cap (budget comes only from FRAME_SAMPLE_RATE × file frames).
MAX_FRAMES_PER_VIDEO = int(os.environ.get("MAX_FRAMES_PER_VIDEO", "0"))
# Fraction of the video stream’s frame count to use for vision (~0.1 = 10%). Requires ffprobe frame estimate.
FRAME_SAMPLE_RATE = float(os.environ.get("FRAME_SAMPLE_RATE", "0.1"))
FRAMES_PER_OLLAMA_BATCH = int(os.environ.get("FRAMES_PER_OLLAMA_BATCH", "4"))
SCENE_CHANGE_EXTRA_SAMPLES = int(os.environ.get("SCENE_CHANGE_EXTRA_SAMPLES", "4"))

OLLAMA_HOST = os.environ.get("OLLAMA_HOST", "http://localhost:11434")
LOCAL_MODEL_NAME = os.environ.get("LOCAL_MODEL_NAME", "qwen2.5vl:7b")
FALLBACK_MODELS = [
    m.strip()
    for m in os.environ.get("FALLBACK_MODELS", "openbmb/minicpm-v4.5,llava:7b").split(",")
    if m.strip()
]

DRY_RUN = os.environ.get("DRY_RUN", "false").lower() in ("1", "true", "yes")
DRY_RUN_ROWS = int(os.environ.get("DRY_RUN_ROWS", "0"))  # 0 = all rows (only if DRY_RUN)
MAX_ROWS_PER_RUN = int(os.environ.get("MAX_ROWS_PER_RUN", "0"))  # 0 = unlimited; use 1 for a test run
START_EXCEL_ROW = int(os.environ.get("START_EXCEL_ROW", "0"))  # 0 = all rows; else skip Excel rows < this
DEBUG_KEEP_VIDEOS = os.environ.get("DEBUG_KEEP_VIDEOS", "false").lower() in ("1", "true", "yes")
FORCE_REPROCESS = os.environ.get("FORCE_REPROCESS", "false").lower() in ("1", "true", "yes")
RESET_DRAFT_WORKBOOK = os.environ.get("RESET_DRAFT_WORKBOOK", "false").lower() in ("1", "true", "yes")

# yt-dlp: age-restricted / private videos need cookies (see README / .env.example).
YTDLP_METADATA_TIMEOUT = int(os.environ.get("YTDLP_METADATA_TIMEOUT", "400"))
YTDLP_TRANSCRIPT_TIMEOUT = int(os.environ.get("YTDLP_TRANSCRIPT_TIMEOUT", "320"))
YTDLP_COOKIES = (os.environ.get("YTDLP_COOKIES") or "").strip()
YTDLP_COOKIES_FROM_BROWSER = (os.environ.get("YTDLP_COOKIES_FROM_BROWSER") or "").strip()
# Mitigates YouTube SABR / "no video formats" with default clients (see yt-dlp #12482). Override with YTDLP_EXTRACTOR_ARGS= to disable.
YTDLP_EXTRACTOR_ARGS = (os.environ.get("YTDLP_EXTRACTOR_ARGS", "youtube:player_client=web,default") or "").strip()

TEMP_DIR = Path(os.environ.get("YT_ANNOTATOR_TEMP", tempfile.gettempdir())) / "corso_youtube_annotator"
TEMP_DIR.mkdir(parents=True, exist_ok=True)

_PROJECT_DIR = Path(__file__).resolve().parent


def _resolve_cookies_file(path_str: str) -> str:
    """Make cookies path work when .env uses a name relative to this project folder."""
    p = Path(path_str).expanduser()
    if p.is_file():
        return str(p.resolve())
    alt = _PROJECT_DIR / path_str
    if alt.is_file():
        return str(alt.resolve())
    return path_str

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("annotate_videos")


def _vision_frame_budget(total_frames_in_file: int | None) -> int:
    """Frames to send to vision: FRAME_SAMPLE_RATE × stream frames, capped by MAX (0 = no cap)."""
    cap = MAX_FRAMES_PER_VIDEO if MAX_FRAMES_PER_VIDEO > 0 else 10**9
    fallback = 24  # when ffprobe cannot estimate frame count
    if total_frames_in_file and total_frames_in_file > 0 and FRAME_SAMPLE_RATE > 0:
        n = max(1, int(round(total_frames_in_file * FRAME_SAMPLE_RATE)))
        return min(n, cap)
    return min(fallback, cap)


def _sample_interval_for_budget(duration_sec: float, budget: int) -> float:
    """Tighten temporal spacing when the vision budget is large so extraction yields enough candidates."""
    if duration_sec <= 0 or budget <= 0:
        return FRAME_INTERVAL_SEC
    need_spacing = duration_sec / max(budget * 1.25, 1.0)
    return max(0.25, min(FRAME_INTERVAL_SEC, need_spacing))


def _fmt_eta(seconds: float) -> str:
    if seconds < 0 or seconds > 86400 * 14 or not (seconds == seconds):
        return "?"
    if seconds < 90:
        return f"{int(seconds)}s"
    m = int(seconds // 60)
    s = int(seconds % 60)
    if m < 60:
        return f"{m}m {s}s"
    h, m = m // 60, m % 60
    return f"{h}h {m}m"


def _parse_rate_frac(s: str | None) -> float | None:
    if not s or s == "0/0":
        return None
    try:
        if "/" in s:
            a, b = s.split("/", 1)
            den = float(b)
            return float(a) / den if den else None
        return float(s)
    except (ValueError, TypeError):
        return None


def _probe_video_duration_and_frames(video_path: Path) -> tuple[float | None, int | None]:
    """
    ffprobe: container duration (seconds) and total frames (exact from nb_frames, else fps * duration).
    """
    try:
        r = _run_cmd(
            [
                "ffprobe",
                "-v",
                "error",
                "-select_streams",
                "v:0",
                "-show_entries",
                "stream=nb_frames,avg_frame_rate,r_frame_rate",
                "-show_entries",
                "format=duration",
                "-of",
                "json",
                str(video_path),
            ],
            timeout=120,
        )
        if r.returncode != 0:
            return None, None
        j = json.loads(r.stdout)
        fmt = j.get("format") or {}
        dur_raw = fmt.get("duration")
        duration_sec = float(dur_raw) if dur_raw not in (None, "", "N/A") else None
        streams = j.get("streams") or []
        if not streams:
            return duration_sec, None
        st = streams[0]
        nb = st.get("nb_frames")
        if nb not in (None, "", "N/A") and str(nb).strip().isdigit():
            return duration_sec, int(nb)
        fps = _parse_rate_frac(st.get("avg_frame_rate")) or _parse_rate_frac(st.get("r_frame_rate"))
        if duration_sec and fps and fps > 0:
            return duration_sec, max(1, int(round(duration_sec * fps)))
        return duration_sec, None
    except (json.JSONDecodeError, TypeError, ValueError, OSError) as e:
        log.debug("ffprobe parse skip: %s", e)
        return None, None


# --- Excel helper column names (added if missing) ---
HELPER_COLUMNS = [
    "video_url",
    "title",
    "description",
    "auto_surgery",
    "auto_type",
    "auto_phase_info",
    "auto_timestamps",
    "auto_crop",
    "auto_notes",
    "auto_confidence",
    "auto_review_status",
]


def build_video_url(youtube_id: str) -> str:
    """Build canonical watch URL from 11-character video id."""
    youtube_id = (youtube_id or "").strip()
    return f"https://www.youtube.com/watch?v={youtube_id}"


_YT_ID_RE = re.compile(r"(?:v=|youtu\.be/|embed/)([0-9A-Za-z_-]{11})")


def extract_youtube_id(text: str) -> str | None:
    """Extract 11-char id from URL or raw id string."""
    if not text or not isinstance(text, str):
        return None
    t = text.strip()
    if len(t) == 11 and re.match(r"^[0-9A-Za-z_-]{11}$", t):
        return t
    m = _YT_ID_RE.search(t)
    if m:
        return m.group(1)
    return None


def _subprocess_env() -> dict[str, str]:
    """Ensure node/deno from Homebrew are visible to yt-dlp’s JS challenge solver."""
    env = os.environ.copy()
    prefix = []
    if sys.platform == "darwin":
        prefix = ["/opt/homebrew/bin", "/opt/homebrew/sbin", "/usr/local/bin"]
    elif sys.platform.startswith("linux"):
        prefix = ["/usr/local/bin"]
    for p in reversed(prefix):
        if p and p not in env.get("PATH", "").split(os.pathsep):
            env["PATH"] = p + os.pathsep + env.get("PATH", "")
    return env


def _run_cmd(cmd: list[str], timeout: int | None = 600) -> subprocess.CompletedProcess:
    log.debug("Running: %s", " ".join(cmd[:8]) + ("..." if len(cmd) > 8 else ""))
    return subprocess.run(
        cmd,
        capture_output=True,
        text=True,
        timeout=timeout,
        check=False,
        env=_subprocess_env(),
    )


def _yt_dlp_base() -> list[str]:
    """yt-dlp executable, or ``python -m yt_dlp`` when the binary is not on PATH."""
    b = (os.environ.get("YTDLP_BIN") or "").strip()
    if b:
        return [b]
    if shutil.which("yt-dlp"):
        return ["yt-dlp"]
    return [sys.executable, "-m", "yt_dlp"]


def _yt_dlp_prefix() -> list[str]:
    """Base command plus optional cookies (file wins over browser)."""
    cmd = _yt_dlp_base()
    if YTDLP_COOKIES:
        cmd.extend(["--cookies", _resolve_cookies_file(YTDLP_COOKIES)])
    elif YTDLP_COOKIES_FROM_BROWSER:
        cmd.extend(["--cookies-from-browser", YTDLP_COOKIES_FROM_BROWSER])
    if YTDLP_EXTRACTOR_ARGS:
        cmd.extend(["--extractor-args", YTDLP_EXTRACTOR_ARGS])
    return cmd


def _yt_dlp_stderr_snippet(proc: subprocess.CompletedProcess | None, limit: int = 900) -> str:
    if proc is None:
        return "yt-dlp timed out"
    msg = (proc.stderr or proc.stdout or "").strip()
    msg = re.sub(r"\s+", " ", msg)
    return msg[:limit]


def fetch_youtube_metadata(url: str) -> dict[str, Any]:
    """
    Title, description, duration, subtitles text (if small) via yt-dlp JSON.
    """
    out: dict[str, Any] = {
        "title": "",
        "description": "",
        "duration": 0,
        "transcript_text": "",
        "raw": {},
        "_yt_dlp_error": "",
    }
    try:
        r = _run_cmd(
            _yt_dlp_prefix()
            + ["--no-warnings", "--no-playlist", "--dump-json", "--skip-download", url],
            timeout=YTDLP_METADATA_TIMEOUT,
        )
    except subprocess.TimeoutExpired:
        log.warning("yt-dlp metadata timed out after %ss", YTDLP_METADATA_TIMEOUT)
        out["_yt_dlp_error"] = f"metadata timeout ({YTDLP_METADATA_TIMEOUT}s)"
        return out
    if r.returncode != 0:
        err = _yt_dlp_stderr_snippet(r, 1200)
        log.warning("yt-dlp metadata failed: %s", err[:600])
        out["_yt_dlp_error"] = err
        return out
    data = None
    for line in r.stdout.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        try:
            data = json.loads(line)
            break
        except json.JSONDecodeError:
            continue
    if data is None:
        log.warning("Could not parse yt-dlp JSON")
        return out
    out["raw"] = data
    out["title"] = data.get("title") or ""
    out["description"] = data.get("description") or ""
    out["duration"] = int(data.get("duration") or 0)
    # Try embedded automatic captions JSON (no extra download) — often empty
    subs = data.get("subtitles") or {}
    auto = data.get("automatic_captions") or {}
    for bucket in (subs, auto):
        for lang, entries in bucket.items():
            if lang.startswith("en") and isinstance(entries, list) and entries:
                # Take first en* track; url may need separate fetch — skip heavy download for v1
                break
    return out


def _fetch_vtt_transcript(url: str, out_dir: Path) -> str:
    """Best-effort English auto subs to text (requires network + yt-dlp)."""
    out_dir.mkdir(parents=True, exist_ok=True)
    pattern = str(out_dir / "subs.%(id)s.%(ext)s")
    try:
        r = _run_cmd(
            _yt_dlp_prefix()
            + [
                "--no-warnings",
                "--no-playlist",
                "--skip-download",
                "--write-auto-subs",
                "--sub-langs",
                "en.*,en",
                "--sub-format",
                "vtt",
                "-o",
                pattern,
                url,
            ],
            timeout=YTDLP_TRANSCRIPT_TIMEOUT,
        )
    except subprocess.TimeoutExpired:
        return ""
    if r.returncode != 0:
        return ""
    texts: list[str] = []
    for p in sorted(out_dir.glob("subs.*.vtt")):
        try:
            raw = p.read_text(encoding="utf-8", errors="ignore")
            # Strip WEBVTT timing lines crudely
            lines = [
                ln.strip()
                for ln in raw.splitlines()
                if ln.strip()
                and not ln.strip().startswith("WEBVTT")
                and "-->" not in ln
                and not re.match(r"^\d+$", ln.strip())
            ]
            texts.append(" ".join(lines[:2000]))  # cap size
        except OSError:
            continue
    return " ".join(texts)[:12000]


def download_smallest_video(url: str, out_dir: Path) -> tuple[Path | None, str]:
    """
    Prefer small combined video+audio suitable for sampling (<= ~480p if available).
    Returns (path, "") on success, or (None, error_message).
    """
    out_dir.mkdir(parents=True, exist_ok=True)
    out_tmpl = str(out_dir / "%(id)s.%(ext)s")
    # Prefer ≤480p merged video+audio; avoid strict ext filters (often no match on YouTube).
    fmt = (
        "bestvideo[height<=480]+bestaudio/"
        "bestvideo[height<=720]+bestaudio/"
        "bestvideo+bestaudio/"
        "best/worst"
    )
    try:
        r = _run_cmd(
            _yt_dlp_prefix()
            + [
                "--no-warnings",
                "--no-playlist",
                "-f",
                fmt,
                "--merge-output-format",
                "mp4",
                "-o",
                out_tmpl,
                url,
            ],
            timeout=3600,
        )
    except subprocess.TimeoutExpired:
        log.error("yt-dlp download timed out")
        return None, "download timed out (3600s)"
    if r.returncode != 0:
        err = _yt_dlp_stderr_snippet(r, 1500)
        log.error("yt-dlp download failed: %s", err[-800:])
        return None, err
    vid = extract_youtube_id(url)
    candidates = list(out_dir.glob(f"{vid}.*")) if vid else list(out_dir.glob("*"))
    videos = [p for p in candidates if p.suffix.lower() in {".mp4", ".webm", ".mkv"}]
    if not videos:
        return None, "no video file written after yt-dlp"
    return max(videos, key=lambda p: p.stat().st_mtime), ""


def _even_sample_indices(count: int, max_n: int) -> list[int]:
    """0..count-1 inclusive, evenly pick at most max_n indices."""
    if count <= 0 or max_n <= 0:
        return []
    if count <= max_n:
        return list(range(count))
    return [int(round(i * (count - 1) / (max_n - 1))) for i in range(max_n)]


def extract_sampled_frames(
    video_path: Path,
    frame_interval_sec: float,
    max_frames: int,
    extra_midpoints: int = 0,
) -> list[tuple[float, Path]]:
    """
    Return list of (timestamp_sec, jpeg_path) using ffmpeg fps filter; fallback OpenCV.
    """
    work = video_path.parent / f"frames_{video_path.stem}"
    if work.exists():
        shutil.rmtree(work)
    work.mkdir(parents=True)

    duration = 0.0
    prob = _run_cmd(
        [
            "ffprobe",
            "-v",
            "error",
            "-show_entries",
            "format=duration",
            "-of",
            "default=noprint_wrappers=1:nokey=1",
            str(video_path),
        ],
        timeout=60,
    )
    if prob.returncode == 0 and prob.stdout.strip():
        try:
            duration = float(prob.stdout.strip())
        except ValueError:
            pass

    fps_expr = f"1/{max(frame_interval_sec, 0.5)}"
    out_pattern = str(work / "f_%04d.jpg")
    r = _run_cmd(
        [
            "ffmpeg",
            "-hide_banner",
            "-loglevel",
            "error",
            "-i",
            str(video_path),
            "-vf",
            f"fps={fps_expr}",
            "-q:v",
            "4",
            out_pattern,
        ],
        timeout=600,
    )
    paths = sorted(work.glob("f_*.jpg"))
    if r.returncode != 0 or not paths:
        log.warning("ffmpeg frame extract failed, using OpenCV fallback")
        paths = _extract_frames_opencv(video_path, work, frame_interval_sec, max_frames)
        # approximate timestamps
        return [(i * frame_interval_sec, p) for i, p in enumerate(paths[:max_frames])]

    # Map frame index to timestamp ~ index * interval
    stamped: list[tuple[float, Path]] = [
        (i * frame_interval_sec, p) for i, p in enumerate(paths)
    ]
    if duration > 0 and extra_midpoints > 0 and len(stamped) < max_frames:
        for k in range(extra_midpoints):
            t = duration * (k + 1) / (extra_midpoints + 1)
            img_path = work / f"extra_{k:04d}.jpg"
            if _ffmpeg_grab_frame(video_path, t, img_path):
                stamped.append((t, img_path))
    stamped.sort(key=lambda x: x[0])
    if len(stamped) > max_frames:
        idxs = _even_sample_indices(len(stamped), max_frames)
        stamped = [stamped[i] for i in idxs]
    return stamped[:max_frames]


def _ffmpeg_grab_frame(video_path: Path, t_sec: float, out_jpg: Path) -> bool:
    r = _run_cmd(
        [
            "ffmpeg",
            "-hide_banner",
            "-loglevel",
            "error",
            "-ss",
            str(max(0.0, t_sec)),
            "-i",
            str(video_path),
            "-frames:v",
            "1",
            "-q:v",
            "4",
            str(out_jpg),
        ],
        timeout=120,
    )
    return r.returncode == 0 and out_jpg.is_file()


def _extract_frames_opencv(
    video_path: Path, work: Path, interval: float, max_frames: int
) -> list[Path]:
    import cv2

    cap = cv2.VideoCapture(str(video_path))
    if not cap.isOpened():
        return []
    fps = cap.get(cv2.CAP_PROP_FPS) or 25.0
    step = max(int(fps * interval), 1)
    paths: list[Path] = []
    frame_i = 0
    saved = 0
    while saved < max_frames:
        cap.set(cv2.CAP_PROP_POS_FRAMES, frame_i)
        ok, frame = cap.read()
        if not ok:
            break
        p = work / f"cv_{saved:04d}.jpg"
        cv2.imwrite(str(p), frame)
        paths.append(p)
        saved += 1
        frame_i += step
    cap.release()
    return paths


def _image_to_b64(path: Path) -> str:
    return base64.b64encode(path.read_bytes()).decode("ascii")


def _chat_vision(
    model: str,
    prompt: str,
    image_paths: list[Path],
    json_only: bool = False,
) -> str:
    """Call Ollama chat with images; return assistant message content."""
    import ollama

    client = ollama.Client(host=OLLAMA_HOST)
    images_b64 = [_image_to_b64(p) for p in image_paths]
    sys_hint = ""
    if json_only:
        sys_hint = (
            "You must respond with valid JSON only. No markdown, no code fences, no commentary."
        )
    messages: list[dict[str, Any]] = []
    if sys_hint:
        messages.append({"role": "system", "content": sys_hint})
    messages.append(
        {
            "role": "user",
            "content": prompt,
            "images": images_b64,
        }
    )
    try:
        resp = client.chat(model=model, messages=messages)
        msg = resp.get("message") or {}
        return (msg.get("content") or "").strip()
    except Exception as e:
        log.exception("Ollama vision call failed: %s", e)
        raise


def _chat_text(model: str, prompt: str, json_only: bool = True) -> str:
    import ollama

    client = ollama.Client(host=OLLAMA_HOST)
    sys_h = (
        "Reply with a single JSON object only. No markdown fences."
        if json_only
        else ""
    )
    messages: list[dict[str, Any]] = []
    if sys_h:
        messages.append({"role": "system", "content": sys_h})
    messages.append({"role": "user", "content": prompt})
    resp = client.chat(model=model, messages=messages)
    return ((resp.get("message") or {}).get("content") or "").strip()


def _extract_json_object(text: str) -> dict[str, Any] | None:
    text = text.strip()
    m = re.search(r"```(?:json)?\s*(\{.*?\})\s*```", text, re.DOTALL)
    if m:
        text = m.group(1)
    try:
        return json.loads(text)
    except json.JSONDecodeError:
        pass
    start = text.find("{")
    end = text.rfind("}")
    if start >= 0 and end > start:
        try:
            return json.loads(text[start : end + 1])
        except json.JSONDecodeError:
            return None
    return None


def analyze_frames_with_ollama(
    frame_batch: list[tuple[float, Path]],
    metadata: dict[str, Any],
    transcript_text: str,
    model_name: str,
) -> list[dict[str, Any]]:
    """
    Vision pass: per-batch strict JSON list of frame analyses.
    """
    rel_times = ", ".join(f"{t:.1f}s" for t, _ in frame_batch)
    paths = [p for _, p in frame_batch]
    title = (metadata.get("title") or "")[:500]
    desc = (metadata.get("description") or "")[:1500]
    tr = (transcript_text or "")[:3000]
    prompt = f"""You are assisting with surgical video screening.

Video title: {title}
Description excerpt: {desc}
Transcript excerpt: {tr}

These images are frames from times: {rel_times} (seconds).

For EACH frame in order, decide if it shows REAL operative surgery footage (patient tissue, operative field, instruments in use) vs non-surgery (talking head, slides, animation, waiting, room setup only).

Return JSON object with exactly one key "frames" whose value is an array of objects, one per frame in order, each with:
- "time_sec": number
- "surgery_visible": true or false
- "view_type": short string: one of open_field, laparoscopic_view, robotic_view, endoscopic_view, non_surgery, unsure
- "confidence": number 0-1
- "notes": short string

Example: {{"frames":[{{"time_sec":0,"surgery_visible":false,"view_type":"non_surgery","confidence":0.9,"notes":""}}]}}"""
    last_err = None
    for attempt in range(2):
        try:
            if attempt == 0:
                raw = _chat_vision(model_name, prompt, paths, json_only=True)
            else:
                raw = _chat_vision(
                    model_name,
                    "Your last reply was not valid JSON. Output ONLY valid JSON: one object with key "
                    '"frames" mapping to an array of per-frame objects. No markdown.',
                    paths,
                    json_only=True,
                )
            data = _extract_json_object(raw)
            if not data or "frames" not in data:
                last_err = "missing frames key"
                continue
            out = []
            for i, (t, _) in enumerate(frame_batch):
                row = data["frames"][i] if i < len(data["frames"]) else {}
                out.append(
                    {
                        "time_sec": float(row.get("time_sec", t)),
                        "surgery_visible": bool(row.get("surgery_visible", False)),
                        "view_type": str(row.get("view_type", "unsure")),
                        "confidence": float(row.get("confidence", 0.5)),
                        "notes": str(row.get("notes", "")),
                    }
                )
            return out
        except Exception as e:
            last_err = str(e)
            log.warning("Frame batch parse failed (%s)", last_err)
    return [
        {
            "time_sec": t,
            "surgery_visible": False,
            "view_type": "unsure",
            "confidence": 0.0,
            "notes": f"model_fail:{last_err}",
        }
        for t, _ in frame_batch
    ]


def merge_surgery_segments(
    frame_predictions: list[dict[str, Any]], duration_sec: float
) -> str:
    """
    Build MM:SS–MM:SS ranges where surgery_visible is true (merge adjacent).
    """
    if not frame_predictions:
        return ""

    def fmt_mmss(sec: float) -> str:
        sec = max(0.0, sec)
        m = int(sec // 60)
        s = int(round(sec % 60))
        return f"{m:02d}:{s:02d}"

    spans: list[tuple[float, float]] = []
    times = sorted(frame_predictions, key=lambda x: x["time_sec"])
    for eff in times:
        if not eff.get("surgery_visible"):
            continue
        t = float(eff["time_sec"])
        if spans and t <= spans[-1][1] + 15:
            spans[-1] = (spans[-1][0], max(spans[-1][1], t + 1))
        else:
            spans.append((t, t + 1))
    parts = [f"{fmt_mmss(a)}–{fmt_mmss(b)}" for a, b in spans]
    return "; ".join(parts)


def aggregate_row_with_ollama(
    metadata: dict[str, Any],
    transcript_text: str,
    frame_summaries: list[dict[str, Any]],
    model_name: str,
) -> dict[str, Any]:
    """Second pass: text-only aggregation → final row JSON."""
    summary_lines = []
    for f in frame_summaries[:80]:
        summary_lines.append(
            f"t={f['time_sec']:.1f}s surgery={f['surgery_visible']} "
            f"view={f['view_type']} conf={f['confidence']:.2f} {f['notes']}"
        )
    block = "\n".join(summary_lines)
    title = metadata.get("title") or ""
    desc = (metadata.get("description") or "")[:4000]
    prompt = f"""You combine metadata and sampled frame analyses for ONE YouTube surgical education video.

Title: {title}
Description:
{desc}

Transcript excerpt:
{(transcript_text or '')[:6000]}

Frame analysis summary:
{block}

Return ONE JSON object with these keys ONLY:
- surgery_or_no_surgery: "Surgery" or "No surgery"
- surgery_type: exactly one of "", "Open", "Laparoscopic", "Robotic", "L/R"
- phase_info: short phrase describing whether phase/step is conveyed via narration, on-screen text, description, single continuous phase, or none
- timestamps: string, ranges in format MM:SS–MM:SS; MM:SS–MM:SS (only where actual surgery is visible; empty string if no surgery)
- crop: "Crop" or "No crop" (Crop if surgery only in PiP/small inset/slide layout)
- notes: short string including endoscopic rule: if no incision and non-invasive, say endoscopic in notes; if unsure lap vs endoscopic say "endoscopic but unsure"
- confidence: number 0-1 (your overall confidence)
- review_status: exactly one of "high_confidence", "medium_confidence", "low_confidence"

Rules:
- Use surgery_type "" when surgery_or_no_surgery is "No surgery".
- Use "L/R" only when unsure laparoscopic vs robotic.

Output JSON only."""

    raw = _chat_text(model_name, prompt, json_only=True)
    data = _extract_json_object(raw)
    if not data:
        raw2 = _chat_text(
            model_name,
            "Return ONLY valid JSON with keys: surgery_or_no_surgery, surgery_type, phase_info, timestamps, crop, notes, confidence, review_status.",
            json_only=True,
        )
        data = _extract_json_object(raw2)
    return data or {}


def _normalize_final(obj: dict[str, Any]) -> dict[str, Any]:
    """Clamp outputs to allowed literals."""
    out = dict(obj)
    s = out.get("surgery_or_no_surgery", "")
    if s not in ("Surgery", "No surgery"):
        out["surgery_or_no_surgery"] = "No surgery" if "no" in str(s).lower() else "Surgery"
    st = out.get("surgery_type", "")
    if st not in ("", "Open", "Laparoscopic", "Robotic", "L/R"):
        out["surgery_type"] = ""
    cr = out.get("crop", "")
    if cr not in ("Crop", "No crop"):
        out["crop"] = "No crop"
    rs = out.get("review_status", "")
    if rs not in ("high_confidence", "medium_confidence", "low_confidence"):
        out["review_status"] = "low_confidence"
    try:
        out["confidence"] = float(out.get("confidence", 0.5))
    except (TypeError, ValueError):
        out["confidence"] = 0.5
    for k in ("phase_info", "timestamps", "notes"):
        out[k] = str(out.get(k, ""))[:8000]
    return out


def _try_models_with_fallback(fn, primary: str, fallbacks: list[str]):
    models = [primary] + [m for m in fallbacks if m != primary]
    last = None
    for m in models:
        try:
            return fn(m), m
        except Exception as e:
            log.warning("Model %s failed: %s", m, e)
            last = e
    raise RuntimeError(f"All models failed: {last}")


# --- Excel ---


def _ensure_output_workbook(input_path: Path, output_path: Path) -> None:
    if output_path.resolve() == input_path.resolve():
        raise ValueError("OUTPUT_XLSX must differ from INPUT_XLSX")
    if RESET_DRAFT_WORKBOOK and output_path.exists():
        output_path.unlink()
        log.info("RESET_DRAFT_WORKBOOK: removed previous draft")
    if not output_path.exists():
        shutil.copy2(input_path, output_path)
        log.info("Created draft workbook: %s", output_path)


def _find_header_row_and_columns(ws) -> tuple[int, dict[str, int]]:
    """Return 1-based header row index and column name -> 1-based col index."""
    for row_idx in range(1, min(ws.max_row, 30) + 1):
        vals = [ws.cell(row_idx, c).value for c in range(1, ws.max_column + 1)]
        non_empty = [str(v).strip() for v in vals if v is not None and str(v).strip()]
        if len(non_empty) >= 3:
            colmap: dict[str, int] = {}
            for c, v in enumerate(vals, start=1):
                if v is None:
                    continue
                key = str(v).strip()
                if key:
                    colmap[key] = c
            return row_idx, colmap
    return 1, {}


def _resolve_youtube_columns(
    colmap: dict[str, int],
) -> tuple[str | None, str | None]:
    """Pick link and/or id column."""
    id_col = YOUTUBE_ID_COLUMN
    link_col = YOUTUBE_LINK_COLUMN
    lower = {k.lower(): k for k in colmap}
    if not link_col:
        for cand in ("link", "url", "youtube", "video link", "youtube url"):
            for k in lower:
                if cand in k:
                    link_col = lower[k]
                    break
            if link_col:
                break
    if not id_col:
        for cand in ("id", "video id", "youtube id"):
            for k in lower:
                if cand == k or k.endswith(cand):
                    id_col = lower[k]
                    break
            if id_col:
                break
    return id_col, link_col


def write_row_results(
    output_path: Path,
    sheet_name: str,
    header_row: int,
    excel_row: int,
    colmap: dict[str, int],
    results: dict[str, str | float],
) -> None:
    import openpyxl

    wb = openpyxl.load_workbook(output_path)
    ws = wb[sheet_name]
    max_c = ws.max_column
    name_to_col = {
        str(ws.cell(header_row, c).value).strip(): c
        for c in range(1, max_c + 1)
        if ws.cell(header_row, c).value
    }

    for name in HELPER_COLUMNS:
        if name not in name_to_col:
            max_c += 1
            ws.cell(header_row, max_c, value=name)
            name_to_col[name] = max_c

    for key, val in results.items():
        if key not in name_to_col:
            max_c += 1
            ws.cell(header_row, max_c, value=key)
            name_to_col[key] = max_c
        ws.cell(excel_row, name_to_col[key], value=val)

    wb.save(output_path)
    wb.close()


def cleanup_temp_files(video_path: Path | None, frame_dirs: list[Path]) -> None:
    for d in frame_dirs:
        if d.exists():
            shutil.rmtree(d, ignore_errors=True)
    if video_path and video_path.exists() and not DEBUG_KEEP_VIDEOS:
        try:
            video_path.unlink()
        except OSError as e:
            log.warning("Could not delete video %s: %s", video_path, e)


def _excel_results_from_ytdlp_failure(
    video_url: str, meta: dict[str, Any], dl_err: str
) -> dict[str, str | float]:
    """When download or metadata fails: still fill the draft row so the run can continue."""
    meta_err = str(meta.get("_yt_dlp_error") or "").strip()
    dle = dl_err.strip()
    parts: list[str] = []
    seen: set[str] = set()
    for p in (meta_err, dle):
        if p and p not in seen:
            seen.add(p)
            parts.append(p)
    note = " | ".join(parts) if parts else "yt-dlp failed"
    low = note.lower()
    if any(
        x in low
        for x in ("sign in", "cookies", "age", "private", "confirm your age", "inappropriate")
    ):
        note += (
            " — Add YTDLP_COOKIES (path to cookies.txt) or YTDLP_COOKIES_FROM_BROWSER "
            "(e.g. chrome, safari, firefox) in .env; see README."
        )
    note = note[:8000]
    return {
        "video_url": video_url,
        "title": meta.get("title", "") or "",
        "description": str(meta.get("description") or "")[:5000],
        "auto_surgery": "No surgery",
        "auto_type": "",
        "auto_phase_info": "none (yt-dlp error)",
        "auto_timestamps": "",
        "auto_crop": "No crop",
        "auto_notes": note,
        "auto_confidence": 0.0,
        "auto_review_status": "low_confidence",
    }


def row_already_done(ws, excel_row: int, colmap: dict[str, int]) -> bool:
    if FORCE_REPROCESS:
        return False
    status_col = colmap.get("auto_review_status")
    if not status_col:
        return False
    v = ws.cell(excel_row, status_col).value
    return bool(v and str(v).strip())


def _pending_youtube_excel_rows(
    ws,
    header_row: int,
    colmap: dict[str, int],
    id_col: str | None,
    link_col: str | None,
) -> list[int]:
    """Excel row indices (1-based) still needing work: YouTube id present, not done, START_EXCEL_ROW."""
    out: list[int] = []
    for excel_row in range(header_row + 1, ws.max_row + 1):
        if START_EXCEL_ROW and excel_row < START_EXCEL_ROW:
            continue
        if row_already_done(ws, excel_row, colmap):
            continue
        raw_id = ws.cell(excel_row, colmap[id_col]).value if id_col else None
        raw_link = ws.cell(excel_row, colmap[link_col]).value if link_col else None
        yt_id = None
        if raw_link:
            yt_id = extract_youtube_id(str(raw_link))
        if not yt_id and raw_id:
            yt_id = extract_youtube_id(str(raw_id))
        if yt_id:
            out.append(excel_row)
    return out


def process_one_row(
    excel_row: int,
    youtube_id: str,
    video_url: str,
    output_path: Path,
    sheet_name: str,
    header_row: int,
    colmap: dict[str, int],
) -> None:
    log.info("Row %s: video %s", excel_row, youtube_id)
    meta: dict[str, Any] = {"title": "", "description": "", "duration": 0}
    if not DRY_RUN:
        meta = fetch_youtube_metadata(video_url)
    sub_dir = TEMP_DIR / youtube_id
    sub_dir.mkdir(parents=True, exist_ok=True)
    transcript = ""
    if not DRY_RUN:
        transcript = _fetch_vtt_transcript(video_url, sub_dir / "subs")

    video_path: Path | None = None
    frame_dirs: list[Path] = []
    final: dict[str, Any] = {}

    try:
        if DRY_RUN:
            log.info("DRY_RUN: skip download and vision")
            final = _normalize_final(
                {
                    "surgery_or_no_surgery": "No surgery",
                    "surgery_type": "",
                    "phase_info": "none (dry run)",
                    "timestamps": "",
                    "crop": "No crop",
                    "notes": "dry run",
                    "confidence": 0.0,
                    "review_status": "low_confidence",
                }
            )
        else:
            video_path, dl_err = download_smallest_video(video_url, sub_dir)
            if not video_path:
                results = _excel_results_from_ytdlp_failure(video_url, meta, dl_err)
                write_row_results(output_path, sheet_name, header_row, excel_row, colmap, results)
                return

            sz_b = video_path.stat().st_size
            sz_mb = sz_b / (1024 * 1024)
            meta_dur = int(meta.get("duration") or 0)
            probe_dur, total_frames_in_file = _probe_video_duration_and_frames(video_path)
            log.info(
                "Downloaded video: %s | %.2f MiB (%s bytes) | yt-dlp metadata duration=%ss",
                video_path.resolve(),
                sz_mb,
                sz_b,
                meta_dur if meta_dur > 0 else "?",
            )
            if probe_dur is not None:
                log.info("ffprobe: duration %.2fs (full merged file)", probe_dur)
            if total_frames_in_file is not None:
                log.info("ffprobe: ~%s frames in video stream (exact or fps×duration)", total_frames_in_file)

            budget = _vision_frame_budget(total_frames_in_file)
            dur_for_interval = (
                probe_dur if probe_dur is not None else float(meta.get("duration") or 0) or 0.0
            )
            interval_eff = _sample_interval_for_budget(dur_for_interval, budget)
            if interval_eff + 1e-6 < FRAME_INTERVAL_SEC:
                log.info(
                    "Sampling interval %.2fs (denser than default %.2fs) for vision budget %s frames (rate=%.0f%%)",
                    interval_eff,
                    FRAME_INTERVAL_SEC,
                    budget,
                    FRAME_SAMPLE_RATE * 100,
                )

            stamped = extract_sampled_frames(
                video_path,
                interval_eff,
                budget,
                extra_midpoints=SCENE_CHANGE_EXTRA_SAMPLES,
            )
            if stamped:
                frame_dirs.append(stamped[0][1].parent)
            else:
                log.warning("No frames extracted from %s", video_path)

            n_sampled = len(stamped)
            from_suffix = (
                f" | subset of ~{total_frames_in_file} frames in file" if total_frames_in_file else ""
            )
            log.info(
                "Sampling: %s frames to vision (budget %s, rate %.0f%%, interval ~%ss)%s",
                n_sampled,
                budget,
                FRAME_SAMPLE_RATE * 100,
                interval_eff,
                from_suffix,
            )

            all_frames: list[dict[str, Any]] = []
            batch_size = FRAMES_PER_OLLAMA_BATCH
            num_batches = (n_sampled + batch_size - 1) // batch_size if n_sampled else 0
            batch_durations: list[float] = []
            for bi, i in enumerate(range(0, len(stamped), batch_size)):
                batch = stamped[i : i + batch_size]
                t_b0 = time.perf_counter()

                def _do_frames(model_name: str):
                    return analyze_frames_with_ollama(
                        batch,
                        {**meta, "title": meta.get("title"), "description": meta.get("description")},
                        transcript,
                        model_name,
                    )

                preds, used_model = _try_models_with_fallback(
                    _do_frames, LOCAL_MODEL_NAME, FALLBACK_MODELS
                )
                batch_durations.append(time.perf_counter() - t_b0)
                done_b = bi + 1
                avg_b = sum(batch_durations) / len(batch_durations)
                rem_b = num_batches - done_b
                eta_batches = avg_b * rem_b if rem_b > 0 else 0.0
                log.info(
                    "Frame vision %s/%s batches | frames %s–%s | %s | ~%s left for vision batches",
                    done_b,
                    max(num_batches, 1),
                    i,
                    i + len(batch) - 1,
                    used_model,
                    _fmt_eta(eta_batches),
                )
                all_frames.extend(preds)

            ts = merge_surgery_segments(all_frames, float(meta.get("duration") or 0))

            def _do_agg(model_name: str):
                a = aggregate_row_with_ollama(
                    meta,
                    transcript,
                    all_frames,
                    model_name,
                )
                a["timestamps"] = a.get("timestamps") or ts
                return _normalize_final(a)

            t_agg0 = time.perf_counter()
            final, used_m = _try_models_with_fallback(_do_agg, LOCAL_MODEL_NAME, FALLBACK_MODELS)
            log.info(
                "Aggregation: model=%s elapsed=%.0fs",
                used_m,
                time.perf_counter() - t_agg0,
            )

        results = {
            "video_url": video_url,
            "title": meta.get("title", ""),
            "description": (meta.get("description") or "")[:5000],
            "auto_surgery": final.get("surgery_or_no_surgery", ""),
            "auto_type": final.get("surgery_type", ""),
            "auto_phase_info": final.get("phase_info", ""),
            "auto_timestamps": final.get("timestamps", ""),
            "auto_crop": final.get("crop", ""),
            "auto_notes": final.get("notes", ""),
            "auto_confidence": final.get("confidence", 0.0),
            "auto_review_status": final.get("review_status", ""),
        }
        write_row_results(output_path, sheet_name, header_row, excel_row, colmap, results)
    finally:
        cleanup_temp_files(video_path, frame_dirs)
        if sub_dir.exists() and not DEBUG_KEEP_VIDEOS:
            shutil.rmtree(sub_dir, ignore_errors=True)


def main() -> int:
    input_path = Path(INPUT_XLSX).expanduser().resolve()
    output_path = Path(OUTPUT_XLSX).expanduser().resolve()

    if not input_path.is_file():
        log.error("Input workbook not found: %s", input_path)
        return 1

    if not YTDLP_COOKIES and not YTDLP_COOKIES_FROM_BROWSER:
        log.warning(
            "YouTube cookies not set (YTDLP_COOKIES / YTDLP_COOKIES_FROM_BROWSER). "
            "Age-restricted and many surgical videos will fail until you export cookies; see README."
        )

    _ensure_output_workbook(input_path, output_path)

    import openpyxl

    wb = openpyxl.load_workbook(output_path, read_only=False, data_only=False)
    sheet = SHEET_NAME if SHEET_NAME in wb.sheetnames else wb.sheetnames[0]
    if sheet not in wb.sheetnames:
        log.error("Sheet %r not found. Available: %s", sheet, wb.sheetnames)
        return 1
    ws = wb[sheet]
    header_row, _ = _find_header_row_and_columns(ws)
    colmap: dict[str, int] = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(header_row, c).value
        if v is not None and str(v).strip():
            colmap[str(v).strip()] = c

    id_col, link_col = _resolve_youtube_columns(colmap)
    log.info("Using id column %s, link column %s", id_col, link_col)

    pending_rows = _pending_youtube_excel_rows(ws, header_row, colmap, id_col, link_col)
    n_pending = len(pending_rows)
    if MAX_ROWS_PER_RUN:
        log.info(
            "Pending YouTube rows in sheet: %s (this run stops after %s successful row(s))",
            n_pending,
            MAX_ROWS_PER_RUN,
        )
    else:
        log.info("Pending YouTube rows in sheet: %s", n_pending)

    processed = 0
    row_wall_times: list[float] = []
    for excel_row in range(header_row + 1, ws.max_row + 1):
        if START_EXCEL_ROW and excel_row < START_EXCEL_ROW:
            continue
        if DRY_RUN_ROWS and processed >= DRY_RUN_ROWS:
            break
        if row_already_done(ws, excel_row, colmap):
            log.info("Skip row %s (already has auto_review_status)", excel_row)
            continue

        raw_id = None
        if id_col:
            raw_id = ws.cell(excel_row, colmap[id_col]).value
        raw_link = None
        if link_col:
            raw_link = ws.cell(excel_row, colmap[link_col]).value

        yt_id = None
        if raw_link:
            yt_id = extract_youtube_id(str(raw_link))
        if not yt_id and raw_id:
            yt_id = extract_youtube_id(str(raw_id))
        if not yt_id:
            log.warning("Row %s: no YouTube id, skipping", excel_row)
            continue

        url = build_video_url(yt_id)
        try:
            t_row0 = time.perf_counter()
            process_one_row(
                excel_row,
                yt_id,
                url,
                output_path,
                sheet,
                header_row,
                colmap,
            )
            processed += 1
            row_wall_times.append(time.perf_counter() - t_row0)
            avg_row = sum(row_wall_times) / len(row_wall_times)
            if MAX_ROWS_PER_RUN:
                left = max(0, MAX_ROWS_PER_RUN - processed)
                eta_run = avg_row * left
                log.info(
                    "Run progress: %s/%s row(s) done | ETA for capped run ~%s (avg %.0fs/row)",
                    processed,
                    MAX_ROWS_PER_RUN,
                    _fmt_eta(eta_run),
                    avg_row,
                )
            else:
                left = sum(1 for r in pending_rows if r > excel_row)
                eta_run = avg_row * left
                log.info(
                    "Run progress: %s row(s) done, ~%s pending after this row | ETA ~%s (avg %.0fs/row)",
                    processed,
                    left,
                    _fmt_eta(eta_run),
                    avg_row,
                )
            if MAX_ROWS_PER_RUN and processed >= MAX_ROWS_PER_RUN:
                log.info("MAX_ROWS_PER_RUN=%s reached; stopping.", MAX_ROWS_PER_RUN)
                break
        except Exception as e:
            log.exception("Row %s unexpected error: %s", excel_row, e)
        wb.close()
        wb = openpyxl.load_workbook(output_path, read_only=False, data_only=False)
        ws = wb[sheet]
        colmap.clear()
        for c in range(1, ws.max_column + 1):
            v = ws.cell(header_row, c).value
            if v is not None and str(v).strip():
                colmap[str(v).strip()] = c

    wb.close()
    log.info("Done. Processed %s rows this run.", processed)
    return 0


if __name__ == "__main__":
    sys.exit(main())
